import os
import re
import logging
import pandas as pd
import dlt
import openpyxl
import xml.etree.ElementTree as ET
import concurrent.futures
from pathlib import Path
from typing import List, Dict, Callable
from ravioli.backend.core.config import settings

logger = logging.getLogger(__name__)

# --- Pipeline Utils ---
def create_ravioli_pipeline(pipeline_name: str, dataset_name: str = "main"):
    """Creates a dlt pipeline configured to use the Ravioli DuckDB instance."""
    return dlt.pipeline(
        pipeline_name=pipeline_name,
        destination=dlt.destinations.duckdb(str(settings.duckdb_path)),
        dataset_name=dataset_name
    )

# --- PII Scanning ---
class PIIScanner:
    """A lightweight, rule-based scanner for PII."""
    PATTERNS = {
        "email": r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}',
        "phone": r'\+?\d{1,4}?[-.\s]?\(?\d{1,3}?\)?[-.\s]?\d{1,4}[-.\s]?\d{1,4}[-.\s]?\d{1,9}',
        "credit_card": r'\b(?:\d[ -]*?){13,16}\b',
        "ssn": r'\b\d{3}-\d{2}-\d{4}\b',
        "ipv4": r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b'
    }

    def __init__(self):
        self.compiled_patterns = {name: re.compile(pattern) for name, pattern in self.PATTERNS.items()}

    def scan_string(self, text: str) -> List[str]:
        if not text or not isinstance(text, str): return []
        found = []
        for name, pattern in self.compiled_patterns.items():
            if pattern.search(text): found.append(name)
        return found

    def scan_dataframe(self, df: pd.DataFrame, sample_size: int = 100) -> bool:
        if df.empty: return False
        sample = df.head(sample_size)
        for column in sample.columns:
            if sample[column].dtype == 'object':
                for value in sample[column].dropna():
                    if self.scan_string(str(value)): return True
        return False

pii_scanner = PIIScanner()

# --- XLSX Processing Utils ---
def process_sheet_with_analysis(df: pd.DataFrame, analysis: dict) -> pd.DataFrame:
    """Apply structural fixes discovered by AI analysis."""
    h_idx = int(analysis.get("header_row", 0))
    d_idx = int(analysis.get("data_start_row", h_idx + 1))
    is_split = analysis.get("is_split", False)
    offsets = analysis.get("split_offsets", [])
    
    if not is_split:
        h = df.iloc[h_idx].tolist()
        seen = set()
        for i, val in enumerate(h):
            v = str(val).lower().strip()
            if v and "unnamed" not in v and v in seen:
                is_split = True
                offsets = [j for j, x in enumerate(h) if str(x).lower().strip() == str(h[0]).lower().strip() and j > 0]
                break
            seen.add(v)
            
    if is_split:
        data = reconcile_split_table(df, h_idx, d_idx, offsets)
    else:
        data = df.iloc[d_idx:].copy()
        data.columns = [str(x).strip() if pd.notna(x) else f"col_{i}" for i, x in enumerate(df.iloc[h_idx])]
        
    data = data.dropna(axis=1, how='all').dropna(axis=0, how='all')
    if analysis.get("column_mapping"):
        data = data.rename(columns={k: v for k, v in analysis["column_mapping"].items() if k in data.columns})
    return data

def reconcile_split_table(df: pd.DataFrame, h_idx: int, d_idx: int, offsets: list) -> pd.DataFrame:
    """Merge multiple side-by-side table blocks into a single vertical table."""
    offsets = sorted(list(set(offsets)))
    blocks = []
    blocks.append(extract_block(df, h_idx, d_idx, 0, offsets[0]))
    for i in range(len(offsets) - 1):
        blocks.append(extract_block(df, h_idx, d_idx, offsets[i], offsets[i+1]))
    blocks.append(extract_block(df, h_idx, d_idx, offsets[-1], df.shape[1]))
    return pd.concat([b for b in blocks if b is not None and not b.empty], ignore_index=True)

def extract_block(df: pd.DataFrame, h_idx: int, d_idx: int, s_col: int, e_col: int) -> pd.DataFrame:
    """Helper to extract and clean a single block from a split table."""
    h = [str(x).strip() if pd.notna(x) else f"col_{i}" for i, x in enumerate(df.iloc[h_idx, s_col:e_col])]
    b = df.iloc[d_idx:, s_col:e_col].copy()
    b.columns = h
    return b.dropna(axis=1, how='all').dropna(axis=0, how='all')

def xlsx_chunk_generator(path: Path, sheet_name: str, analysis: dict, chunk_size: int = 50000):
    """Memory-efficient streaming generator for massive XLSX sheets."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    h_idx = int(analysis.get("header_row", 0))
    d_idx = int(analysis.get("data_start_row", h_idx + 1))
    
    # Get headers
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == h_idx:
            header_row = [str(x).strip() if x is not None else f"col_{j}" for j, x in enumerate(row)]
            break
            
    if not header_row: return

    chunk = []
    total_processed = 0
    for i, row in enumerate(ws.iter_rows(min_row=d_idx + 1, values_only=True)):
        # Convert row to dict
        record = dict(zip(header_row, row))
        chunk.append(record)
        total_processed += 1
        if len(chunk) >= chunk_size:
            logger.info(f"Chunking: Processed {total_processed:,} rows from sheet '{sheet_name}'")
            yield chunk
            chunk = []
            
    if chunk:
        logger.info(f"Chunking complete for sheet '{sheet_name}'. Total: {total_processed:,} rows.")
        yield chunk

# --- XML Ingestion Utils ---
import mmap
import re

def xml_tag_generator(path: Path, tag_name: str, extract_metadata: bool = False):
    """Generic generator to stream specific XML tags using iterparse."""
    context = ET.iterparse(path, events=("end",))
    for _, elem in context:
        if elem.tag == tag_name:
            data = dict(elem.attrib)
            if extract_metadata:
                metadata = {c.attrib.get('key'): c.attrib.get('value') for c in elem if c.tag == "MetadataEntry"}
                if metadata: data['metadata'] = metadata
            yield data
        elem.clear()

def xml_chunk_generator(path: Path, tag_name: str, start: int, end: int, extract_metadata: bool = False):
    """Memory-efficient generator using mmap to scan XML chunks without loading them into RAM."""
    attr_pattern = re.compile(r'(\w+)="([^"]*)"')
    count = 0
    
    with open(path, "rb") as f:
        with mmap.mmap(f.fileno(), length=0, access=mmap.ACCESS_READ) as mm:
            view = mm[start:end]
            
            if tag_name == "Record":
                pattern = re.compile(rf'<{tag_name}\s+([^>]+)\s*/>'.encode())
                for match in pattern.finditer(view):
                    attrs_raw = match.group(1).decode('utf-8', errors='ignore')
                    count += 1
                    if count % 100000 == 0:
                        logger.info(f"Chunk [{start//1024**2}MB]: Found {count:,} records...")
                    yield dict(attr_pattern.findall(attrs_raw))
            else:
                pattern = re.compile(rf'<{tag_name}\s+([^>/]+)\s*(?:/>|>(.*?)</{tag_name}>)'.encode(), re.DOTALL)
                for match in pattern.finditer(view):
                    attrs_raw = match.group(1).decode('utf-8', errors='ignore')
                    entry = dict(attr_pattern.findall(attrs_raw))
                    if extract_metadata and match.group(2):
                        inner = match.group(2).decode('utf-8', errors='ignore')
                        meta_pattern = re.compile(r'<MetadataEntry\s+key="([^"]*)"\s+value="([^"]*)"\s*/>')
                        metadata = {k: v for k, v in meta_pattern.findall(inner)}
                        if metadata: entry['metadata'] = metadata
                    count += 1
                    if count % 100000 == 0:
                        logger.info(f"Chunk [{start//1024**2}MB]: Found {count:,} records...")
                    yield entry
            
            logger.info(f"Chunk completed. Total found: {count:,} records.")
            del view

def xml_full_parse_generator(path: Path, original_filename: str):
    """Fallback generator for full XML parsing."""
    tree = ET.parse(path)
    root = tree.getroot()
    yield {"filename": original_filename, "root_tag": root.tag, "attribs": dict(root.attrib)}

def scan_xml_chunk(path: Path, tag_name: str, start: int, end: int, extract_metadata: bool):
    """Worker function for parallel XML scanning."""
    results = []
    attr_pattern = re.compile(r'(\w+)="([^"]*)"')
    
    with open(path, 'rb') as f:
        f.seek(start)
        data = f.read(end - start)
        
        if tag_name == "Record":
            # optimized for flat Record tags
            tag_pattern = re.compile(rf'<{tag_name}\s+([^>]+)\s*/>'.encode())
            for match in tag_pattern.finditer(data):
                attrs_raw = match.group(1).decode('utf-8', errors='ignore')
                results.append(dict(attr_pattern.findall(attrs_raw)))
        elif tag_name == "Workout":
            # handle Workout tags with potential MetadataEntry children
            tag_pattern = re.compile(rf'<{tag_name}\s+([^>]+)>(.*?)</{tag_name}>'.encode(), re.DOTALL)
            for match in tag_pattern.finditer(data):
                attrs_raw = match.group(1).decode('utf-8', errors='ignore')
                entry = dict(attr_pattern.findall(attrs_raw))
                if extract_metadata:
                    inner = match.group(2).decode('utf-8', errors='ignore')
                    meta_pattern = re.compile(r'<MetadataEntry\s+key="([^"]*)"\s+value="([^"]*)"\s*/>')
                    metadata = {k: v for k, v in meta_pattern.findall(inner)}
                    if metadata: entry['metadata'] = metadata
                results.append(entry)
        else:
            # generic fallback for other tags
            tag_pattern = re.compile(rf'<{tag_name}\s+([^>/]+)\s*(?:/>|>(.*?)</{tag_name}>)'.encode(), re.DOTALL)
            for match in tag_pattern.finditer(data):
                attrs_raw = match.group(1).decode('utf-8', errors='ignore')
                results.append(dict(attr_pattern.findall(attrs_raw)))
    return results

def parallel_xml_tag_generator(path: Path, tag_name: str, extract_metadata: bool = False, num_workers: int = 4):
    """Heavylift: Parallelized XML scanning for massive files."""
    size = os.path.getsize(path)
    chunk_size = size // num_workers
    chunks = []
    
    with open(path, 'rb') as f:
        for i in range(num_workers):
            start = i * chunk_size
            if i > 0:
                f.seek(start)
                f.readline() # align to next line
                start = f.tell()
            
            end = (i + 1) * chunk_size if i < num_workers - 1 else size
            if i < num_workers - 1:
                f.seek(end)
                f.readline() # align to end of line
                end = f.tell()
            chunks.append((start, end))

    logger.info(f"Starting parallel XML ingestion with {num_workers} workers for tag {tag_name}")
    with concurrent.futures.ProcessPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(scan_xml_chunk, path, tag_name, start, end, extract_metadata): (start, end) for start, end in chunks}
        completed_workers = 0
        for future in concurrent.futures.as_completed(futures):
            completed_workers += 1
            start, end = futures[future]
            logger.info(f"Worker {completed_workers}/{num_workers} finished chunk [{start:,} - {end:,}] for tag {tag_name}")
            try:
                results = future.result()
                logger.info(f"Chunk yielded {len(results):,} records.")
                yield from results
            except Exception as e:
                logger.error(f"Worker failed for chunk [{start} - {end}]: {e}")

XML_STRATEGIES = {
    "apple_health_export": {
        "match": lambda fn: 'export' in fn and 'cda' not in fn,
        "tables": [
            {"tag": "Record", "table_name": "apple_health_records"},
            {"tag": "Workout", "table_name": "apple_health_workouts", "extract_metadata": True},
            {"tag": "ActivitySummary", "table_name": "apple_health_activity_summaries"}
        ]
    },
    "apple_health_cda": {
        "match": lambda fn: 'export_cda' in fn,
        "tables": [
            {"tag": "ClinicalRecord", "table_name": "apple_health_clinical_records"}
        ]
    }
}
